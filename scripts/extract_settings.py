#!/usr/bin/env python3
"""
Extract run settings from Excel CREST model Main Sheet.

This script reads simulation parameters directly from the Excel Main Sheet
and outputs them in various formats for use with Python simulations.

How it works:
1. Opens Excel file using openpyxl (reads calculated values only)
2. Extracts settings from specific cells in Main Sheet:
   - F6: Day of month
   - H6: Month of year
   - F7: Weekday/Weekend
   - F8: Latitude, H8: Longitude, L8: Meridian
   - F9: City/location
   - F10: Country
   - H10: Year (for India simulations)
   - K10: Urban/Rural setting
   - F12: Number of dwellings
3. Reads checkbox values from VML XML inside the XLSM file:
   - objAssignDwellingParameters: Stochastically assign dwelling parameters
   - objDynamicOutput: Include high-resolution dynamic output
   - objDailyTotals: Include daily demand totals
   - objOverwriteData: Overwrite existing data
   - objDaylightSaving: Country uses daylight saving time
   - objPVOption: PV included as an option
4. Outputs in requested format (text, json, or shell script)

Usage:
    python scripts/extract_settings.py excel/original.xlsm
    python scripts/extract_settings.py excel/original.xlsm --format json
    python scripts/extract_settings.py excel/original.xlsm --format shell
    python scripts/extract_settings.py excel/original.xlsm --output settings.json
"""

import sys
import argparse
import json
import zipfile
import re
from pathlib import Path
from typing import Dict, Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)


def extract_checkbox_values(excel_path: Path) -> Dict[str, bool]:
    """
    Extract checkbox values from VML XML inside the XLSM file.

    XLSM files are ZIP archives. The checkbox controls are stored in
    xl/drawings/vmlDrawing1.vml as XML with shape IDs matching the
    checkbox names (e.g., objAssignDwellingParameters).

    Args:
        excel_path: Path to .xlsm file

    Returns:
        Dictionary mapping checkbox name to checked state (True/False)
    """
    checkboxes = {}

    try:
        with zipfile.ZipFile(excel_path, 'r') as z:
            # Find all vmlDrawing files
            vml_files = [name for name in z.namelist() if 'vmlDrawing' in name]

            for vml_file in vml_files:
                content = z.read(vml_file).decode('utf-8')

                # Parse checkbox shapes using regex
                # Shape format: <v:shape id="objName" ...>...<x:Checked>1</x:Checked>...</v:shape>
                # Pattern to find shape blocks with checkbox ObjectType
                shape_pattern = r'<v:shape\s+id="([^"]+)"[^>]*>.*?</v:shape>'

                for match in re.finditer(shape_pattern, content, re.DOTALL):
                    shape_block = match.group(0)
                    shape_id = match.group(1)

                    # Check if this is a checkbox
                    if 'ObjectType="Checkbox"' in shape_block:
                        # Check if it's checked
                        # <x:Checked>1</x:Checked> means checked
                        # No <x:Checked> tag or <x:Checked>0</x:Checked> means unchecked
                        checked_match = re.search(r'<x:Checked>(\d+)</x:Checked>', shape_block)
                        is_checked = checked_match is not None and checked_match.group(1) == '1'

                        # Only store checkboxes with meaningful names (not _x0000_s####)
                        if not shape_id.startswith('_x0000_'):
                            checkboxes[shape_id] = is_checked

    except Exception as e:
        print(f"  Warning: Could not read checkbox values: {e}")

    return checkboxes


def extract_settings(excel_path: Path) -> Dict[str, Any]:
    """
    Extract run settings from Main Sheet.

    Settings locations (based on original Excel layout):
    - Column H contains most parameters (starting around row 6)
    - K10: Additional parameter
    - M8: Additional parameter
    - Checkboxes need special handling

    Args:
        excel_path: Path to .xlsm file

    Returns:
        Dictionary of settings
    """
    wb = load_workbook(excel_path, data_only=True)

    # Try to find the Main Sheet (handle variations in naming)
    main_sheet = None
    for sheet_name in wb.sheetnames:
        if 'main' in sheet_name.lower():
            main_sheet = wb[sheet_name]
            break

    if main_sheet is None:
        raise ValueError(f"Could not find 'Main Sheet' in {excel_path}. Available sheets: {wb.sheetnames}")

    print(f"Reading settings from sheet: '{main_sheet.title}'")

    settings = {}

    # Extract settings based on actual Main Sheet layout
    # Row 6: Date - Day in F6, Month in H6
    # Row 7: Weekday/Weekend in F7
    # Row 8: Latitude in F8, Longitude in H8, LST Meridian in L8
    # Row 9: City in F9
    # Row 10: Country in F10, Year in H10, Urban/Rural in K10
    # Row 12: Number of dwellings in F12

    # Day of month (F6)
    day_val = main_sheet['F6'].value
    settings['day'] = int(day_val) if day_val is not None else 1

    # Month of year (H6)
    month_val = main_sheet['H6'].value
    settings['month'] = int(month_val) if month_val is not None else 1

    # Weekday/Weekend (F7)
    weekday_val = main_sheet['F7'].value
    settings['weekday'] = str(weekday_val).lower() if weekday_val is not None else 'wd'

    # Latitude (F8)
    lat_val = main_sheet['F8'].value
    if lat_val is not None:
        try:
            settings['latitude'] = float(lat_val)
        except (ValueError, TypeError):
            settings['latitude'] = 52.77  # Default: UK
    else:
        settings['latitude'] = 52.77

    # Longitude (H8)
    lon_val = main_sheet['H8'].value
    if lon_val is not None:
        try:
            settings['longitude'] = float(lon_val)
        except (ValueError, TypeError):
            settings['longitude'] = -1.26  # Default: UK
    else:
        settings['longitude'] = -1.26

    # LST Meridian (M8) - note: label is in J8, value is in M8
    meridian_val = main_sheet['M8'].value
    if meridian_val is not None:
        try:
            settings['meridian'] = float(meridian_val)
        except (ValueError, TypeError):
            settings['meridian'] = 0.0  # Default: UK/England
    else:
        settings['meridian'] = 0.0

    # City/location (F9)
    city_val = main_sheet['F9'].value
    settings['city'] = str(city_val) if city_val is not None else 'England'

    # Country (F10)
    country_val = main_sheet['F10'].value
    settings['country'] = str(country_val) if country_val is not None else 'UK'

    # Year (H10) - for India simulations
    year_val = main_sheet['H10'].value
    if year_val is not None:
        try:
            settings['year'] = int(float(year_val))
        except (ValueError, TypeError):
            settings['year'] = 2006
    else:
        settings['year'] = 2006

    # Urban/Rural (K10)
    urban_rural_val = main_sheet['K10'].value
    settings['urban_rural'] = str(urban_rural_val) if urban_rural_val is not None else 'Urban'

    # Number of dwellings (F12)
    num_dwellings_val = main_sheet['F12'].value
    settings['num_dwellings'] = int(num_dwellings_val) if num_dwellings_val is not None else 1

    # Seed - not typically in Main Sheet, default to None
    settings['seed'] = None

    # Extract checkbox values from VML XML
    checkboxes = extract_checkbox_values(excel_path)

    # Map checkbox names to settings
    # objAssignDwellingParameters: Stochastically assign dwelling parameters (row 13)
    settings['assign_dwelling_params'] = checkboxes.get('objAssignDwellingParameters', True)

    # objDynamicOutput: Include high-resolution dynamic output (row 14)
    settings['save_detailed'] = checkboxes.get('objDynamicOutput', True)

    # objDailyTotals: Include daily demand totals (row 15)
    settings['save_daily_totals'] = checkboxes.get('objDailyTotals', True)

    # objOverwriteData: Overwrite existing data (row 16)
    settings['overwrite_data'] = checkboxes.get('objOverwriteData', True)

    # objPVOption: PV included as an option (row 17)
    settings['pv_enabled'] = checkboxes.get('objPVOption', True)

    # objDaylightSaving: Country uses daylight saving time (row 11)
    settings['daylight_saving'] = checkboxes.get('objDaylightSaving', True)

    # Portable RNG is a Python-specific setting, not in Excel
    settings['use_portable_rng'] = False

    # Print what we found for debugging
    print("\nExtracted settings:")
    for key, value in sorted(settings.items()):
        print(f"  {key:20} = {value}")

    return settings


def format_as_shell_script(settings: Dict[str, Any], excel_path: Path) -> str:
    """
    Format settings as a shell script for re-running the simulation.

    The script includes all extracted settings in an easy-to-read format,
    allowing exact replication of the Excel configuration in Python.

    Args:
        settings: Settings dictionary
        excel_path: Original Excel file path

    Returns:
        Shell script content
    """
    from datetime import datetime

    script_lines = [
        "#!/bin/bash",
        "# CREST Simulation Run Script",
        f"# Generated from: {excel_path}",
        f"# Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "# ============================================================",
        "# SIMULATION SETTINGS (extracted from Excel Main Sheet)",
        "# ============================================================",
        "",
        "# Date settings",
        f"DAY={settings.get('day', 1)}",
        f"MONTH={settings.get('month', 1)}",
        f"WEEKDAY=\"{settings.get('weekday', 'wd')}\"  # 'wd' = weekday, 'we' = weekend",
        "",
        "# Location settings",
        f"LATITUDE={settings.get('latitude', 52.77)}",
        f"LONGITUDE={settings.get('longitude', -1.26)}",
        f"MERIDIAN={settings.get('meridian', 0.0)}  # Local standard time meridian",
        f"CITY=\"{settings.get('city', 'England')}\"",
        f"COUNTRY=\"{settings.get('country', 'UK')}\"",
        f"YEAR={settings.get('year', 2006)}",
        f"URBAN_RURAL=\"{settings.get('urban_rural', 'Urban')}\"",
        "",
        "# Simulation settings",
        f"NUM_DWELLINGS={settings.get('num_dwellings', 1)}",
        f"SEED={settings.get('seed') if settings.get('seed') is not None else '\"\"'}  # Empty for random seed",
        "",
        "# Checkbox settings (from Excel form controls)",
        f"ASSIGN_DWELLING_PARAMS={'true' if settings.get('assign_dwelling_params', True) else 'false'}  # Stochastically assign dwelling parameters",
        f"SAVE_DETAILED={'true' if settings.get('save_detailed', True) else 'false'}  # Include high-resolution dynamic output",
        f"SAVE_DAILY_TOTALS={'true' if settings.get('save_daily_totals', True) else 'false'}  # Include daily demand totals",
        f"OVERWRITE_DATA={'true' if settings.get('overwrite_data', True) else 'false'}  # Overwrite existing data",
        f"PV_ENABLED={'true' if settings.get('pv_enabled', True) else 'false'}  # PV included as an option",
        f"DAYLIGHT_SAVING={'true' if settings.get('daylight_saving', True) else 'false'}  # Country uses daylight saving time",
        "",
        "# Python-specific settings",
        f"USE_PORTABLE_RNG={'true' if settings.get('use_portable_rng', False) else 'false'}  # Use portable LCG for RNG validation",
        "",
        "# ============================================================",
        "# PATHS (set these before running)",
        "# ============================================================",
        "",
        "# Config file containing dwelling configurations",
        "DWELLINGS_FILE=\"${DWELLINGS_FILE:-excel/lcg_fixed/Dwellings.csv}\"",
        "",
        "# Output directory for results",
        "OUTPUT_DIR=\"${OUTPUT_DIR:-output/run}\"",
        "",
        "# ============================================================",
        "# RUN THE SIMULATION",
        "# ============================================================",
        "",
        "# Build command line arguments",
        "CMD_ARGS=()",
        "CMD_ARGS+=(--day \"$DAY\")",
        "CMD_ARGS+=(--month \"$MONTH\")",
        "CMD_ARGS+=(--latitude \"$LATITUDE\")",
        "CMD_ARGS+=(--longitude \"$LONGITUDE\")",
        "CMD_ARGS+=(--meridian \"$MERIDIAN\")",
        "CMD_ARGS+=(--country \"$COUNTRY\")",
        "CMD_ARGS+=(--city \"$CITY\")",
        "CMD_ARGS+=(--year \"$YEAR\")",
        "CMD_ARGS+=(--urban-rural \"$URBAN_RURAL\")",
        "CMD_ARGS+=(--config-file \"$DWELLINGS_FILE\")",
        "CMD_ARGS+=(--output-dir \"$OUTPUT_DIR\")",
        "",
        "# Add seed if specified",
        "if [ -n \"$SEED\" ]; then",
        "    CMD_ARGS+=(--seed \"$SEED\")",
        "fi",
        "",
        "# Add optional flags based on checkbox settings",
        "if [ \"$SAVE_DETAILED\" = \"true\" ]; then",
        "    CMD_ARGS+=(--save-detailed)",
        "fi",
        "",
        "if [ \"$USE_PORTABLE_RNG\" = \"true\" ]; then",
        "    CMD_ARGS+=(--portable-rng)",
        "fi",
        "",
        "# Run the simulation",
        "echo \"Running CREST simulation with settings from: ${excel_path}\"",
        "echo \"Output directory: $OUTPUT_DIR\"",
        "echo \"\"",
        "",
        "venv/bin/python python/main.py \"${CMD_ARGS[@]}\"",
        "",
    ]

    return '\n'.join(script_lines)


def format_as_json(settings: Dict[str, Any]) -> str:
    """Format settings as JSON."""
    return json.dumps(settings, indent=2)


def main():
    parser = argparse.ArgumentParser(description='Extract run settings from Excel CREST model')
    parser.add_argument('excel_file', type=Path, help='Path to .xlsm Excel file')
    parser.add_argument('--format', choices=['json', 'shell', 'text'], default='text',
                       help='Output format (default: text)')
    parser.add_argument('--output', type=Path, help='Output file (default: stdout)')

    args = parser.parse_args()

    if not args.excel_file.exists():
        print(f"ERROR: File not found: {args.excel_file}")
        sys.exit(1)

    # Extract settings
    try:
        settings = extract_settings(args.excel_file)
    except Exception as e:
        print(f"ERROR: Failed to extract settings: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    # Format output
    if args.format == 'json':
        output = format_as_json(settings)
    elif args.format == 'shell':
        output = format_as_shell_script(settings, args.excel_file)
    else:  # text
        output = "\n".join([f"{k}={v}" for k, v in sorted(settings.items())])

    # Write output
    if args.output:
        with open(args.output, 'w') as f:
            f.write(output)
        print(f"\nSettings written to: {args.output}")
    else:
        print("\n" + "=" * 60)
        print(output)
        print("=" * 60)


if __name__ == '__main__':
    main()
