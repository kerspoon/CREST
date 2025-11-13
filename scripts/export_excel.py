#!/usr/bin/env python3
"""
Export VBA code and CSV sheets from Excel .xlsm file.

This script extracts VBA macros and CSV data from Excel files to enable
version control and comparison with Python implementation.

How it works:
1. Deletes existing output directory to ensure fresh export
2. Uses oletools.olevba to extract VBA code (modules, classes, forms)
3. Uses openpyxl to export all worksheets as CSV files
4. Saves VBA files with original names (.bas, .cls, .frm extensions)
5. Exits with error if required packages are missing

Requirements:
- openpyxl: Read Excel files and export sheets to CSV
- pandas: DataFrame handling for CSV export
- oletools: Extract VBA macros from .xlsm files (REQUIRED)

Usage:
    python scripts/export_excel.py excel/original.xlsm
    python scripts/export_excel.py excel/original.xlsm --output excel/original/
    python scripts/export_excel.py excel/original.xlsm --vba-only
    python scripts/export_excel.py excel/original.xlsm --csv-only
"""

import sys
import argparse
import shutil
from pathlib import Path

try:
    from openpyxl import load_workbook
    import pandas as pd
    from oletools.olevba import VBA_Parser
except ImportError as e:
    print("ERROR: Required packages not installed.")
    print("Install with: pip install openpyxl pandas oletools")
    print(f"Missing: {e}")
    sys.exit(1)


def export_vba_code(excel_path: Path, output_dir: Path) -> int:
    """
    Export VBA code from Excel file.

    Uses oletools to extract VBA macros from .xlsm files and saves
    each module/class to a separate file.

    Args:
        excel_path: Path to .xlsm file
        output_dir: Directory to save VBA files

    Returns:
        Number of files exported
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        # Parse VBA project
        vbaparser = VBA_Parser(str(excel_path))

        if not vbaparser.detect_vba_macros():
            print(f"  No VBA macros found in {excel_path.name}")
            vbaparser.close()
            return 0

        # Extract all modules
        count = 0
        for (filename, stream_path, vba_filename, vba_code) in vbaparser.extract_macros():
            if vba_code:
                # Determine file extension based on module type
                # VBA modules typically end with .bas, .cls, or .frm
                if vba_filename:
                    # Use the original VBA filename
                    output_file = output_dir / vba_filename
                else:
                    # Generate a filename
                    output_file = output_dir / f"Module{count + 1}.bas"

                # Write VBA code to file
                output_file.write_text(vba_code, encoding='utf-8')
                print(f"  Exported: {output_file.name}")
                count += 1

        vbaparser.close()

        if count == 0:
            print(f"  No VBA modules extracted from {excel_path.name}")

        return count

    except Exception as e:
        print(f"  ERROR extracting VBA: {e}")
        return 0


def export_csv_sheets(excel_path: Path, output_dir: Path) -> int:
    """
    Export all sheets from Excel file as CSV.

    Args:
        excel_path: Path to .xlsm file
        output_dir: Directory to save CSV files

    Returns:
        Number of sheets exported
    """
    csv_dir = output_dir
    csv_dir.mkdir(parents=True, exist_ok=True)

    print(f"  Loading workbook: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)

    count = 0
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Convert to pandas DataFrame
        data = sheet.values
        cols = next(data)  # First row is header
        df = pd.DataFrame(data, columns=cols)

        # Save as CSV
        csv_file = csv_dir / f"{sheet_name}.csv"
        df.to_csv(csv_file, index=False)

        print(f"  Exported: {sheet_name} → {csv_file.name}")
        count += 1

    return count


def main():
    parser = argparse.ArgumentParser(
        description='Export VBA code and CSV sheets from Excel file'
    )
    parser.add_argument('excel_file', type=Path,
                       help='Path to .xlsm Excel file')
    parser.add_argument('--output', '-o', type=Path,
                       help='Output directory (default: excel/{basename}/)')
    parser.add_argument('--vba-only', action='store_true',
                       help='Export only VBA code')
    parser.add_argument('--csv-only', action='store_true',
                       help='Export only CSV sheets')

    args = parser.parse_args()

    if not args.excel_file.exists():
        print(f"ERROR: File not found: {args.excel_file}")
        sys.exit(1)

    # Determine output directory
    if args.output:
        output_dir = args.output
    else:
        # Default: excel/{basename}/
        basename = args.excel_file.stem  # filename without extension
        output_dir = Path("excel") / basename

    # Delete output directory if it exists to start fresh
    if output_dir.exists():
        print(f"Deleting existing directory: {output_dir}")
        shutil.rmtree(output_dir)

    # Create fresh output directory
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Exporting from: {args.excel_file}")
    print(f"Output directory: {output_dir}")
    print()

    # Export VBA (unless --csv-only)
    if not args.csv_only:
        print("Exporting VBA code...")
        export_vba_code(args.excel_file, output_dir)
        print()

    # Export CSV sheets (unless --vba-only)
    if not args.vba_only:
        print("Exporting CSV sheets...")
        num_sheets = export_csv_sheets(args.excel_file, output_dir)
        print(f"\nExported {num_sheets} sheets to CSV")

    print()
    print(f"✓ Export complete: {output_dir}")


if __name__ == '__main__':
    main()
